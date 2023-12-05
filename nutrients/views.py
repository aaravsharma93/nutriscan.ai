from django.contrib import auth
from django.db.models.query import QuerySet
from django.http.response import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from rest_framework.authtoken.models import Token
# from rest_framework.authentication import BasicAuthentication
# from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import authentication_classes, permission_classes, api_view
from nutrients.models import DailyAllowanceSheet, UserTable, UserFamilyMember
from django.db.models import Q
from django.conf import settings

from numpy import floor
import openpyxl
import numpy as np
import os
import json
import nutrients.image_ocr as image_ocr


def index(request):
    if request.method == 'GET':
        nutrients_data = DailyAllowanceSheet.objects.all()
        return render(request, 'nutrients/index.html', {"excel_data": nutrients_data})


def upload_excel_file(request):
    excel_file = request.FILES["excel_file"]
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb.active

    data = parse_excel_file(worksheet)
    uploaded_data = upload_excel_data(data)
    return HttpResponse("Data has been successfully copied!")


def parse_excel_file(worksheet):
    def get_data(row, col):
        try:
            return (
                worksheet.cell(row=1, column=col).value,
                worksheet.cell(row=row, column=col).value
            )
        except:
            return (None, None)
    row_data = list()
    for row in range(2, worksheet.max_row+1):
        excel_data = dict()
        for cell in range(1, worksheet.max_column+1):
            key, value = get_data(row, cell)
            excel_data.update({
                key: value
            })  
        row_data.append(excel_data)
    return row_data


def upload_excel_data(data):
    for row in data:
        content = DailyAllowanceSheet.objects.create(**row)
    return content  


def registerRequest(request):
    username_response = {"Error": "Username already taken!", 'status':404}
    email_response = {"Error": "User already registered for this email!", 'status':404}
    
    params = request.POST
    username = params.get('username')
    firstName = params.get('firstname')
    lastName = params.get('lastname')
    email = params.get('email')
    password = params.get('password')
    
    auth_email = UserTable.objects.filter(email=email)
    auth_username = UserTable.objects.filter(username=username)
    if not auth_email:
        if not auth_username:
            user = UserTable(username=username, email=email, first_name=firstName, last_name=lastName, password=password)
            user.save()
            email = UserTable.objects.filter(email=email)
            idx = email[0].id
            return JsonResponse({"Success": "User Successfully Registered", "status": 200 ,"user_id":idx})
        else:
            return JsonResponse(username_response)
    else:
        return JsonResponse(email_response)       



def loginRequest(request):
    login_response = {"Error": "Inavalid Username or Password!", 'status': 404}
    

    params = request.POST
    user = params.get('username')
    password = params.get('password')
    # return JsonResponse({"Error": "Inavalid Username or Password!", 'status':200 , 'username':user,'password':password})
    usernames = UserTable.objects.filter(Q(username=user) | Q(email=user))
    if usernames.count() == 1:
        if usernames[0].password == password:
            # token = Token.objects.create(user=usernames)
            # print(token.key)
            userid = usernames[0].id
            success_response = {"Success": "LoggedIn Successfully", 'status': 200 ,'user_id':userid }
            return JsonResponse(success_response) 
    else:
        return JsonResponse(login_response)        


def logoutRequest(request):
    logout_response = {'status': 200, 'message': 'Logout Successfully.'}
    return JsonResponse(logout_response)


def family_member(request):
    success_response = {"Success": "Member Added", "status": 200}
    data = request.POST
    member_name = data.get('member_name')
    age = int(data.get('age'))
    gender = data.get('gender')
    
    user_id = data.get('id')
    user_get_id = UserTable.objects.get(id=user_id)
    
    preg = (data.get('pregnant')).title()   
    brfeed = (data.get('breastfeeding')).title()
    # print("preg",preg)
    # print("brfeed",brfeed)
    Queryset = UserFamilyMember.objects.filter(Q(member_name=member_name) & Q(user=user_id) & Q(age=age))
    # if len(Queryset) > 0:
    #     return JsonResponse({'Error':'Member already Exists'})
    # else:
    Queryset = UserFamilyMember(user=user_get_id, member_name=member_name, age=age, gender=gender, pregnant=preg, breastfeeding=brfeed)
    Queryset.save()
    return JsonResponse(success_response)

def getmemberdetails(request):
    data = request.POST
    user_id  = data.get('id')
    print(user_id)
    Queryset = UserFamilyMember.objects.filter(Q(user=user_id))
    member_list = []
    for each in Queryset:
        member_dict = {"id":each.id,
        "name":each.member_name,
        "age":each.age,
        "gender":each.gender,
        "pregnant":each.pregnant,
        "breastfeeing":each.breastfeeding}
        member_list.append(member_dict)
    member_list.reverse()
    success_response = {"Success": "Member Detail Fetched", "status": 200,'member_data':member_list}   
    return JsonResponse(success_response)
    
def edit_members(request):
    data = request.POST
    
    member_id = data.get('id')
    name = data.get('member_name')
    age = data.get('age')
    gender = data.get('gender')
    pregnant = (data.get('pregnant')).title()
    breastfeeding = (data.get('breastfeeding')).title()
    # QuerySet = UserFamilyMember.objects.get(pk=mem

    print(name)
    QuerySet = UserFamilyMember.objects.filter(id=member_id)

    
    if len(QuerySet) == 0:
        return JsonResponse({"Result":"Member Doesnt Exists","code":404})

    QuerySet = UserFamilyMember.objects.get(pk=member_id)
    QuerySet.member_name = name
    QuerySet.age = age
    QuerySet.gender = gender
    QuerySet.pregnant = pregnant
    QuerySet.breastfeeding = breastfeeding

    QuerySet.save()
    success_response = {"Result": "Member Detail Updated Successfully.", "status": 200}   
    return JsonResponse(success_response)


def delete_member(request):
    data = request.POST
    member_id = data.get('id')
    QuerySet = UserFamilyMember.objects.filter(id=member_id)
    # print(type(len(QuerySet)))
    if len(QuerySet) == 0:
        return JsonResponse({"Result":"Member Doesnt Exists","code":404})
    QuerySet = UserFamilyMember.objects.get(pk=member_id)
    QuerySet.delete()


    success_response = {"Result": "Member Successfully Deleted.", "status": 200}   
    return JsonResponse(success_response)






def getDailyAllowance(json_data, img_save_path, pregnant=False, breastfeeding=False):
    # breakpoint()
    details = image_ocr.ocr(img_save_path)
    # print("OCR----",details)
    
    if "Error" in details.keys():
        details['status'] = 404
        return details
    else:
        params = json_data
        age, gender = float(params.get('age')), params.get('gender')
        age_grp = {
            0: '0_6m',
            1: '6m_1',
            2: '1_3',
            3: '4_8',
            4: '9_13',
            5: '14_18',
            6: '19_50',
            7: '50+'
        }

        if age <= 0.5:
            age_index = 0
        elif 0.5 < age <= 1:
            age_index = 1
        elif 1 < age <= 3:
            age_index = 2
        elif 3 < age <= 8:
            age_index = 3
        elif 8 < age <= 13:
            age_index = 4
        elif 13 < age <= 18:
            age_index = 5
        elif 18 < age <= 50:
            age_index = 6
        elif 50 < age:
            age_index = 7

        if gender in ['M', 'Male', 'm', 'male']:
            col = f"m_{age_grp[age_index]}"
        elif gender in ['F', 'Female', 'f', 'female']:
            if bool(params.get('pregnant')):
                col = 'pregnant'
            elif bool(params.get('breastfeeding')):
                col = 'breastfeeding'
            else:
                col = f"f_{age_grp[age_index]}"

        for nutrientName, data in details.items():
            try:
                n = DailyAllowanceSheet.objects.get(nutrient_name=nutrientName)
                # print(n.__dict__)
                if len(n.__dict__) < 1:
                    data['daily_intake_value'] = np.nan
                    data['daily_intake_value_unit'] = None
                else:
                    data['daily_intake_value'] = n.__dict__[col]
                    data['daily_intake_value_unit'] = n.__dict__['unit']
                    # return JsonResponse({nutrientName:n.__dict__[col]})
                details[nutrientName] = data
            except Exception as e:
                data['daily_intake_value'] = np.nan
                data['daily_intake_value_unit'] = None
                print(nutrientName, e)

        normal, excess = 0, 0
        # print(details)
        for item, values in details.items():
            if float(values['value']) < float(values['daily_intake_value']):
                normal += 1
            elif float(values['value']) > float(values['daily_intake_value']):
                excess += 1

        output = {}
        output['status'] = 200
        output['Nutrients_Detected'] = len(details)
        output['normal'] = normal
        output['excess'] = excess
        output['Id'] = json_data.get('Id')
        if len(details) > 0:
            output['Score'] = str(floor((normal/len(details))*100)) + '%'
        else:
            output['Score'] = '0%'
        print(output, '\n')
        # os.remove(img_save_path)

        return output,details


# @api_view(['POST'])
# @authentication_classes([BasicAuthentication])
# @permission_classes([IsAuthenticated])
def picture_Upload(request):
    data = {'status': 404, 'Error': 'Try Again'}
    if request.method == "POST":
        print(request.POST)
        if request.FILES.get("image", None) is not None:
            #So this would be the logic
            img = request.FILES["image"]
            img_name = request.POST.get('image_name')

            # Create image save path with title
            img_save_path = os.path.join(settings.MEDIA_ROOT, img_name)
            with open(img_save_path, "wb+") as f:
                for chunk in img.chunks():
                    f.write(chunk)
            request_id = request.POST.get('id')
            print("ID....",request_id)
            user = UserFamilyMember.objects.filter(user=request_id)
            print("User",user)
            if user:
                member_data = []
                for i in user:
                    json_data = {
                        'Id':i.id,
                        'member_name': i.member_name,
                        'age': i.age,
                        'gender': i.gender,
                        'pregnant': i.pregnant,
                        'breastfeeding': i.breastfeeding
                    }
                    data,nutri_data = getDailyAllowance(json_data, img_save_path)
                    if data["status"] == 404:
                        return JsonResponse({"result":data,"code":400,"nutri_data":[]}) 
                    else:
                        member_data.append({"Id":i.id,"Data":data,"member_data":json_data})
            else:
                return JsonResponse({"result":"User Doesn't Exists","code":400})          
        else:
            print("No file")
            return JsonResponse(data)
    return JsonResponse({"result":member_data,"code":200,"nutri_data":nutri_data})

def copyright(request):
    return render(request, 'nutrients/copyright.html')